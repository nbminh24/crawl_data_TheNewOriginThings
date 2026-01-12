import os
import re
import time
from playwright.sync_api import sync_playwright
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import requests
from io import BytesIO
from dotenv import load_dotenv
import signal
from functools import wraps

load_dotenv()

cloudinary_url = os.getenv('CLOUDINARY_URL')
if cloudinary_url:
    match = re.match(r'cloudinary://([^:]+):([^@]+)@(.+)', cloudinary_url)
    if match:
        api_key, api_secret, cloud_name = match.groups()
        cloudinary.config(
            cloud_name=cloud_name,
            api_key=api_key,
            api_secret=api_secret
        )

class CoolmateCrawler:
    def __init__(self, collection_urls):
        self.collection_urls = collection_urls if isinstance(collection_urls, list) else [collection_urls]
        self.products_data = []
        self.excel_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'lecas_data.xlsx')
        self.wb = None
        self.ws = None
        self.row_index = 2
        
    
    def extract_category(self, url):
        match = re.search(r'/collection/([^/?]+)', url)
        return match.group(1) if match else 'unknown'
    
    def crawl_collection(self, page, collection_url):
        print(f"\nCrawling collection: {collection_url}")
        page.goto(collection_url, wait_until='domcontentloaded', timeout=60000)
        time.sleep(3)
        
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(2)
        
        product_links = page.evaluate("""
            () => {
                const links = Array.from(document.querySelectorAll('a[href*="/product/"]'));
                return [...new Set(links.map(a => a.href))];
            }
        """)
        
        print(f"Found {len(product_links)} products in this collection")
        return product_links
    
    def upload_to_cloudinary(self, image_url, folder_name, timeout=30):
        try:
            result = cloudinary.uploader.upload(
                image_url,
                folder=f"coolmate/{folder_name}",
                use_filename=True,
                unique_filename=True,
                timeout=timeout
            )
            return result['secure_url']
        except Exception as e:
            print(f"⚠️ Upload failed: {str(e)[:100]}")
            return None
    
    
    def get_product_colors(self, page):
        try:
            colors = page.evaluate("""
                () => {
                    const colors = [];
                    const seen = new Set();
                    
                    const colorImgs = document.querySelectorAll('img[alt^="color "]');
                    colorImgs.forEach(img => {
                        const alt = img.getAttribute('alt');
                        if (alt && alt.startsWith('color ')) {
                            const colorName = alt.replace('color ', '').trim();
                            const normalized = colorName.toLowerCase();
                            if (!seen.has(normalized) && colorName.length > 0) {
                                seen.add(normalized);
                                colors.push({name: colorName});
                            }
                        }
                    });
                    
                    if (colors.length === 0) {
                        const urlParams = new URLSearchParams(window.location.search);
                        const currentColor = urlParams.get('color');
                        if (currentColor) {
                            colors.push({name: currentColor, isCurrent: true});
                        }
                    }
                    
                    return colors;
                }
            """)
            return colors if colors and len(colors) > 0 else [{'name': 'default', 'isCurrent': True}]
        except:
            return [{'name': 'default', 'isCurrent': True}]
    
    def crawl_product_detail(self, page, product_url, category):
        print(f"\nCrawling product: {product_url}")
        try:
            page.goto(product_url, wait_until='domcontentloaded', timeout=45000)
            time.sleep(2)
        except Exception as e:
            print(f"⚠️ Failed to load product page: {str(e)[:50]}")
            return
        
        product_name = page.evaluate("""
            () => {
                const h1 = document.querySelector('h1');
                const title = document.querySelector('[class*="product-title"], [class*="ProductTitle"]');
                return h1?.textContent.trim() || title?.textContent.trim() || 'Unknown Product';
            }
        """)
        
        price = page.evaluate("""
            () => {
                const priceEl = document.querySelector('[class*="price"], [class*="Price"], .product-price');
                return priceEl?.textContent.trim() || 'N/A';
            }
        """)
        
        colors = self.get_product_colors(page)
        print(f"Product: {product_name}, Price: {price}, Colors found: {len(colors)}")
        
        for idx, color_info in enumerate(colors, 1):
            color_name = color_info.get('name', 'default')
            print(f"  [{idx}/{len(colors)}] Color: {color_name}")
            
            try:
                if idx > 1:
                    try:
                        print(f"    Clicking color button...", end=' ')
                        clicked = page.evaluate(f"""
                            () => {{
                                const colorImg = document.querySelector('img[alt="color {color_name}"]');
                                if (colorImg) {{
                                    const button = colorImg.closest('button');
                                    if (button) {{
                                        button.click();
                                        return true;
                                    }}
                                }}
                                return false;
                            }}
                        """)
                        
                        if clicked:
                            time.sleep(2)
                            print("✓")
                        else:
                            print("✗ Button not found")
                    except Exception as e:
                        print(f"✗ Failed to click: {str(e)[:50]}")
                
                images = page.evaluate("""
                    () => {
                        const imgs = [];
                        
                        const galleryContainer = document.querySelector('.no-scrollbar.absolute.left-5, [class*="no-scrollbar"]');
                        if (galleryContainer) {
                            const buttons = galleryContainer.querySelectorAll('button img');
                            buttons.forEach(img => {
                                const alt = img.getAttribute('alt');
                                if (!alt || !alt.startsWith('color ')) {
                                    const src = img.src || img.getAttribute('data-src');
                                    if (src && src.includes('n7media.coolmate.me')) {
                                        const cleanSrc = src.split('?')[0];
                                        imgs.push(cleanSrc);
                                    }
                                }
                            });
                        }
                        
                        if (imgs.length === 0) {
                            const allButtons = document.querySelectorAll('button img[alt*="Áo"], button img[alt*="Quần"]');
                            allButtons.forEach(img => {
                                const alt = img.getAttribute('alt');
                                if (!alt || !alt.startsWith('color ')) {
                                    const src = img.src || img.getAttribute('data-src');
                                    if (src && src.includes('n7media.coolmate.me') && src.includes('uploads')) {
                                        const parent = img.closest('.header, .footer, .menu, nav');
                                        if (!parent) {
                                            const cleanSrc = src.split('?')[0];
                                            imgs.push(cleanSrc);
                                        }
                                    }
                                }
                            });
                        }
                        
                        return [...new Set(imgs)];
                    }
                """)
                
                description = page.evaluate("""
                () => {
                    const sections = [];
                    
                    const features = document.querySelectorAll('[class*="feature"], [class*="benefit"], [class*="detail"]');
                    features.forEach(f => {
                        const text = f.textContent.trim();
                        if (text && text.length < 200) sections.push(text);
                    });
                    
                    const details = document.querySelector('[class*="description"], [class*="Detail"], [class*="info"]');
                    if (details) {
                        const lines = details.textContent.split('\\n').map(l => l.trim()).filter(l => l);
                        sections.push(...lines);
                    }
                    
                    return [...new Set(sections)].join('\\n\\n');
                }
                """)
                
                print(f"    Found {len(images)} images")
                
                uploaded_images = []
                max_images = min(len(images), 10)
                for img_idx, img_url in enumerate(images[:max_images]):
                    if img_url.startswith('//'):
                        img_url = 'https:' + img_url
                    elif not img_url.startswith('http'):
                        img_url = 'https:' + img_url if img_url.startswith('//') else None
                    
                    if not img_url:
                        continue
                        
                    print(f"    [{img_idx+1}/{max_images}] Uploading...", end=' ')
                    uploaded_url = self.upload_to_cloudinary(img_url, f"{category}/{product_name.replace(' ', '_')}/{color_name}", timeout=30)
                    if uploaded_url:
                        uploaded_images.append(uploaded_url)
                        print("✓")
                    else:
                        print("✗ Skip")
                
                if len(uploaded_images) > 0:
                    product_data = {
                        'category': category,
                        'product_name': product_name,
                        'price': price,
                        'color': color_name,
                        'images': ', '.join(uploaded_images),
                        'description': description
                    }
                    self.products_data.append(product_data)
                    
                    if self.append_to_excel(product_data):
                        print(f"    ✓ Saved {len(uploaded_images)} images for {color_name} → Excel updated")
                    else:
                        print(f"    ✓ Saved {len(uploaded_images)} images for {color_name} (Excel update failed)")
                else:
                    print(f"    ⚠️ No images saved for {color_name}")
                    
            except Exception as e:
                print(f"    ✗ Error processing color {color_name}: {str(e)[:100]}")
                continue
    
    def init_excel(self):
        from datetime import datetime
        
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Products"
            
            headers = ['STT', 'Category', 'Tên sản phẩm', 'Giá', 'Màu sắc', 'Danh sách link ảnh', 'Mô tả sản phẩm']
            self.ws.append(headers)
            
            for col in self.ws[1]:
                col.font = Font(bold=True)
                col.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                self.ws.column_dimensions[col].width = 20
            self.ws.column_dimensions['F'].width = 80
            self.ws.column_dimensions['G'].width = 50
            
            self.wb.save(self.excel_path)
            print(f"✓ Excel file created: {self.excel_path}\n")
        except PermissionError:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.excel_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'lecas_data_{timestamp}.xlsx')
            print(f"⚠️ File locked, using new file: {self.excel_path}")
            
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Products"
            
            headers = ['STT', 'Category', 'Tên sản phẩm', 'Giá', 'Màu sắc', 'Danh sách link ảnh', 'Mô tả sản phẩm']
            self.ws.append(headers)
            
            for col in self.ws[1]:
                col.font = Font(bold=True)
                col.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                self.ws.column_dimensions[col].width = 20
            self.ws.column_dimensions['F'].width = 80
            self.ws.column_dimensions['G'].width = 50
            
            self.wb.save(self.excel_path)
            print(f"✓ Excel file created: {self.excel_path}\n")
    
    def append_to_excel(self, product_data):
        try:
            self.ws.append([
                self.row_index - 1,
                product_data['category'],
                product_data['product_name'],
                product_data['price'],
                product_data['color'],
                product_data['images'],
                product_data['description']
            ])
            self.row_index += 1
            self.wb.save(self.excel_path)
            return True
        except Exception as e:
            print(f"    ⚠️ Failed to save to Excel: {e}")
            return False
    
    def finalize_excel(self):
        if self.wb:
            self.wb.save(self.excel_path)
            print(f"\n{'='*60}")
            print(f"✓ Final data saved to: {self.excel_path}")
            print(f"Total rows saved: {self.row_index - 2}")
            print(f"{'='*60}")
    
    def run(self):
        print(f"Total collections to crawl: {len(self.collection_urls)}\n")
        
        self.init_excel()
        
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=False)
                page = browser.new_page()
                
                all_products = []
                for idx, collection_url in enumerate(self.collection_urls, 1):
                    print(f"\n{'='*60}")
                    print(f"[Collection {idx}/{len(self.collection_urls)}]")
                    category = self.extract_category(collection_url)
                    print(f"Category: {category}")
                    
                    try:
                        product_links = self.crawl_collection(page, collection_url)
                        for product in product_links:
                            all_products.append((product, category))
                    except Exception as e:
                        print(f"Error crawling collection {collection_url}: {e}")
                        continue
                
                print(f"\n{'='*60}")
                print(f"Total products found: {len(all_products)}")
                print(f"{'='*60}\n")
                
                for idx, (product_url, category) in enumerate(all_products, 1):
                    print(f"\n{'='*60}")
                    print(f"[Product {idx}/{len(all_products)}]")
                    print(f"Progress: {len(self.products_data)} variants saved so far")
                    print(f"URL: {product_url}")
                    
                    try:
                        self.crawl_product_detail(page, product_url, category)
                    except KeyboardInterrupt:
                        raise
                    except Exception as e:
                        print(f"⚠️ Error crawling product: {str(e)[:100]}")
                        print("→ Skipping to next product...")
                        continue
                
                browser.close()
        except KeyboardInterrupt:
            print("\n\n" + "="*60)
            print("⚠️ SCRIPT INTERRUPTED BY USER (Ctrl+C)")
            print(f"Data saved: {len(self.products_data)} variants")
            print("="*60)
        except Exception as e:
            print(f"\n\n⚠️ Script error: {e}")
        finally:
            self.finalize_excel()

if __name__ == "__main__":
    print("=== COOLMATE CRAWLER ===\n")
    print("Nhập danh sách collection URLs (mỗi URL 1 dòng, nhấn Enter 2 lần để kết thúc):")
    print("Ví dụ: https://www.coolmate.me/collection/ao-ba-lo-tank-top-nam\n")
    
    collection_urls = []
    while True:
        url = input().strip()
        if not url:
            break
        if '/collection/' in url:
            collection_urls.append(url)
        else:
            print("⚠️  URL không hợp lệ (phải chứa '/collection/')")
    
    if not collection_urls:
        print("Không có URL nào được nhập. Thoát.")
        exit()
    
    print(f"\nSẽ crawl {len(collection_urls)} collection(s)")
    
    crawler = CoolmateCrawler(collection_urls)
    crawler.run()
