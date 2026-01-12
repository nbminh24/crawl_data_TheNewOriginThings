import os
import re
import time
from playwright.sync_api import sync_playwright
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from dotenv import load_dotenv

load_dotenv()

cloudinary_url = os.getenv('CLOUDINARY_URL')
if cloudinary_url:
    match = re.match(r'cloudinary://([^:]+):([^@]+)@(.+)', cloudinary_url)
    if match:
        api_key, api_secret, cloud_name = match.groups()
        cloudinary.config(
            cloud_name=cloud_name,
            api_key=api_key,
            api_secret=api_secret
        )

class TheNewOriginalsCrawler:
    def __init__(self, collection_urls):
        self.collection_urls = collection_urls if isinstance(collection_urls, list) else [collection_urls]
        self.products_data = []
        self.excel_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'tno_data.xlsx')
        self.wb = None
        self.ws = None
        self.row_index = 2
        self.crawled_products = set()
        
    def extract_category(self, url):
        match = re.search(r'/collections/([^/?]+)', url)
        return match.group(1) if match else 'unknown'
    
    def upload_to_cloudinary(self, image_url, folder_name, timeout=30):
        try:
            if not image_url.startswith('http'):
                image_url = 'https:' + image_url if image_url.startswith('//') else 'https://theneworiginals.co' + image_url
            
            result = cloudinary.uploader.upload(
                image_url,
                folder=f"theneworiginals/{folder_name}",
                use_filename=True,
                unique_filename=True,
                timeout=timeout
            )
            return result['secure_url']
        except Exception as e:
            print(f"⚠️ Upload failed: {str(e)[:100]}")
            return None
    
    def crawl_collection(self, page, collection_url, max_pages=25):
        print(f"\nCrawling collection: {collection_url} (max {max_pages} pages)")
        
        all_products = set()
        current_page = 1
        
        while True:
            if current_page > max_pages:
                print(f"  Reached page limit ({max_pages})")
                break
            
            page_url = f"{collection_url}?page={current_page}" if current_page > 1 else collection_url
            print(f"  Page {current_page}/{max_pages}...", end=' ')
            
            try:
                page.goto(page_url, wait_until='domcontentloaded', timeout=60000)
                time.sleep(2)
                
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(1)
                
                products_on_page = page.evaluate("""
                    () => {
                        const links = Array.from(document.querySelectorAll('a[href*="/products/"]'));
                        const baseUrls = links.map(a => a.href.split('?')[0]);
                        return [...new Set(baseUrls)];
                    }
                """)
                
                if not products_on_page or len(products_on_page) == 0:
                    print("No products found")
                    break
                
                before_count = len(all_products)
                for product in products_on_page:
                    all_products.add(product)
                new_products = len(all_products) - before_count
                
                print(f"{new_products} new products")
                
                has_next = page.evaluate("""
                    () => {
                        const nextButton = document.querySelector('.pagination__item--next:not(.pagination__item--disable)');
                        return nextButton !== null;
                    }
                """)
                
                if not has_next:
                    break
                
                current_page += 1
                
            except Exception as e:
                print(f"Error on page {current_page}: {str(e)[:50]}")
                break
        
        product_links = list(all_products)
        print(f"\nTotal unique products: {len(product_links)}")
        return product_links
    
    def get_all_colors(self, page):
        try:
            colors = page.evaluate("""
                () => {
                    const colorNames = [];
                    
                    const colorInputs = document.querySelectorAll('input[name="Màu"]');
                    colorInputs.forEach(input => {
                        const colorValue = input.value;
                        if (colorValue) {
                            colorNames.push(colorValue);
                        }
                    });
                    
                    if (colorNames.length === 0) {
                        const currentColor = document.querySelector('.current-option[data-selected-value]');
                        if (currentColor) {
                            colorNames.push(currentColor.textContent.trim());
                        }
                    }
                    
                    return colorNames;
                }
            """)
            return colors if colors and len(colors) > 0 else ['N/A']
        except:
            return ['N/A']
    
    def crawl_product_detail(self, page, product_url, category):
        print(f"\nCrawling product: {product_url}")
        try:
            page.goto(product_url, wait_until='domcontentloaded', timeout=45000)
            time.sleep(2)
        except Exception as e:
            print(f"⚠️ Failed to load product page: {str(e)[:50]}")
            return
        
        product_name = page.evaluate("""
            () => {
                const h1 = document.querySelector('h1, .product-title, [class*="product-name"]');
                return h1?.textContent.trim() || 'Unknown Product';
            }
        """)
        
        if product_name in self.crawled_products:
            print(f"  ⏭️  Skipped (already crawled)")
            return
        
        product_name_original = product_name
        
        price = page.evaluate("""
            () => {
                const priceEl = document.querySelector('.price, [class*="price"], .product-price');
                return priceEl?.textContent.trim() || 'N/A';
            }
        """)
        
        colors = self.get_all_colors(page)
        colors_str = ', '.join(colors)
        print(f"Product: {product_name_original}, Price: {price}, Colors: {colors_str}")
                
        images = page.evaluate("""
            () => {
                const imgs = [];
                
                const productImages = document.querySelectorAll('.product-image img, .product-gallery img, [class*="ProductImage"] img, .product__media img');
                productImages.forEach(img => {
                    const src = img.src || img.getAttribute('data-src') || img.getAttribute('srcset')?.split(' ')[0];
                    if (src && !src.includes('icon') && !src.includes('logo')) {
                        const cleanSrc = src.split('?')[0];
                        imgs.push(cleanSrc);
                    }
                });
                
                if (imgs.length === 0) {
                    const allImgs = document.querySelectorAll('img');
                    allImgs.forEach(img => {
                        const parent = img.closest('.header, .footer, .nav, nav, .menu');
                        if (!parent) {
                            const src = img.src || img.getAttribute('data-src');
                            if (src && src.includes('theneworiginals') && !src.includes('icon') && !src.includes('logo')) {
                                const cleanSrc = src.split('?')[0];
                                imgs.push(cleanSrc);
                            }
                        }
                    });
                }
                
                return [...new Set(imgs)];
            }
        """)
                
        description = page.evaluate("""
            () => {
                const sections = [];
                
                const productLabels = document.querySelectorAll('.product-labels__title, .product-labels__description');
                productLabels.forEach(el => {
                    const text = el.textContent.trim();
                    if (text && text.length > 5 && text.length < 300) {
                        sections.push(text);
                    }
                });
                
                const descBlock = document.querySelector('.description-block__text .rte');
                if (descBlock) {
                    const lines = descBlock.textContent.split('\\n').map(l => l.trim()).filter(l => l && l.length > 5);
                    sections.push(...lines);
                }
                
                const accordions = document.querySelectorAll('.accordion__text');
                accordions.forEach(acc => {
                    const text = acc.textContent.trim();
                    if (text && text.length > 10 && text.length < 500) {
                        sections.push(text);
                    }
                });
                
                return [...new Set(sections)].join('\\n\\n');
            }
        """)
                
        print(f"  Found {len(images)} images")
        
        uploaded_images = []
        max_images = min(len(images), 15)
        for img_idx, img_url in enumerate(images[:max_images]):
            print(f"  [{img_idx+1}/{max_images}] Uploading...", end=' ')
            uploaded_url = self.upload_to_cloudinary(img_url, f"{category}/{product_name.replace(' ', '_')}", timeout=30)
            if uploaded_url:
                uploaded_images.append(uploaded_url)
                print("✓")
            else:
                print("✗ Skip")
        
        if len(uploaded_images) > 0:
            product_data = {
                'category': category,
                'product_name': product_name_original,
                'price': price,
                'colors': colors_str,
                'images': ', '.join(uploaded_images),
                'description': description
            }
            self.products_data.append(product_data)
            self.crawled_products.add(product_name_original)
            
            if self.append_to_excel(product_data):
                print(f"  ✓ Saved {len(uploaded_images)} images → Excel updated")
            else:
                print(f"  ✓ Saved {len(uploaded_images)} images (Excel update failed)")
        else:
            print(f"  ⚠️ No images saved")
    
    def load_existing_products(self):
        if os.path.exists(self.excel_path):
            try:
                wb = load_workbook(self.excel_path)
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[2]:  # Tên sản phẩm
                        self.crawled_products.add(row[2])
                
                print(f"ℹ️  Found existing Excel file with {len(self.crawled_products)} products")
                print(f"   Will skip already crawled products\n")
                wb.close()
            except Exception as e:
                print(f"⚠️  Could not read existing Excel: {str(e)[:50]}\n")
    
    def init_excel(self):
        from datetime import datetime
        
        if os.path.exists(self.excel_path):
            try:
                self.wb = load_workbook(self.excel_path)
                self.ws = self.wb.active
                self.row_index = self.ws.max_row + 1
                print(f"✓ Excel file opened (continuing from row {self.row_index})\n")
                return
            except Exception as e:
                print(f"⚠️  Could not open existing Excel: {str(e)[:50]}")
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                self.excel_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'tno_data_{timestamp}.xlsx')
                print(f"   Creating new file: {self.excel_path}\n")
        
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Products"
            
            headers = ['STT', 'Category', 'Tên sản phẩm', 'Giá', 'Màu sắc', 'Danh sách link ảnh', 'Mô tả sản phẩm']
            self.ws.append(headers)
            
            for col in self.ws[1]:
                col.font = Font(bold=True)
                col.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                self.ws.column_dimensions[col].width = 20
            self.ws.column_dimensions['F'].width = 80
            self.ws.column_dimensions['G'].width = 50
            
            self.wb.save(self.excel_path)
            print(f"✓ Excel file created: {self.excel_path}\n")
        except PermissionError:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.excel_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'tno_data_{timestamp}.xlsx')
            print(f"⚠️ File locked, using new file: {self.excel_path}")
            
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Products"
            
            headers = ['STT', 'Category', 'Tên sản phẩm', 'Giá', 'Màu sắc', 'Danh sách link ảnh', 'Mô tả sản phẩm']
            self.ws.append(headers)
            
            for col in self.ws[1]:
                col.font = Font(bold=True)
                col.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                self.ws.column_dimensions[col].width = 20
            self.ws.column_dimensions['F'].width = 80
            self.ws.column_dimensions['G'].width = 50
            
            self.wb.save(self.excel_path)
            print(f"✓ Excel file created: {self.excel_path}\n")
    
    def append_to_excel(self, product_data):
        try:
            self.ws.append([
                self.row_index - 1,
                product_data['category'],
                product_data['product_name'],
                product_data['price'],
                product_data['colors'],
                product_data['images'],
                product_data['description']
            ])
            self.row_index += 1
            self.wb.save(self.excel_path)
            return True
        except Exception as e:
            print(f"    ⚠️ Failed to save to Excel: {e}")
            return False
    
    def finalize_excel(self):
        if self.wb:
            self.wb.save(self.excel_path)
            print(f"\n{'='*60}")
            print(f"✓ Final data saved to: {self.excel_path}")
            print(f"Total rows saved: {self.row_index - 2}")
            print(f"{'='*60}")
    
    def run(self):
        print(f"Total collections to crawl: {len(self.collection_urls)}\n")
        
        self.load_existing_products()
        self.init_excel()
        
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=False)
                page = browser.new_page()
                
                all_products = []
                for idx, collection_url in enumerate(self.collection_urls, 1):
                    print(f"\n{'='*60}")
                    print(f"[Collection {idx}/{len(self.collection_urls)}]")
                    category = self.extract_category(collection_url)
                    print(f"Category: {category}")
                    
                    try:
                        product_links = self.crawl_collection(page, collection_url)
                        for product in product_links:
                            all_products.append((product, category))
                    except Exception as e:
                        print(f"Error crawling collection {collection_url}: {e}")
                        continue
                
                print(f"\n{'='*60}")
                print(f"Total products found: {len(all_products)}")
                print(f"{'='*60}\n")
                
                for idx, (product_url, category) in enumerate(all_products, 1):
                    print(f"\n{'='*60}")
                    print(f"[Product {idx}/{len(all_products)}]")
                    print(f"Progress: {len(self.products_data)} variants saved so far")
                    print(f"URL: {product_url}")
                    
                    try:
                        self.crawl_product_detail(page, product_url, category)
                    except KeyboardInterrupt:
                        raise
                    except Exception as e:
                        print(f"⚠️ Error crawling product: {str(e)[:100]}")
                        print("→ Skipping to next product...")
                        continue
                
                browser.close()
        except KeyboardInterrupt:
            print("\n\n" + "="*60)
            print("⚠️ SCRIPT INTERRUPTED BY USER (Ctrl+C)")
            print(f"Data saved: {len(self.products_data)} variants")
            print("="*60)
        except Exception as e:
            print(f"\n\n⚠️ Script error: {e}")
        finally:
            self.finalize_excel()

if __name__ == "__main__":
    print("=== THE NEW ORIGINALS CRAWLER ===\n")
    print("Nhập danh sách collection URLs (mỗi URL 1 dòng, nhấn Enter 2 lần để kết thúc):")
    print("Ví dụ: https://theneworiginals.co/collections/ao-thun-relaxed-fit\n")
    
    collection_urls = []
    while True:
        url = input().strip()
        if not url:
            break
        if '/collections/' in url:
            collection_urls.append(url)
        else:
            print("⚠️  URL không hợp lệ (phải chứa '/collections/')")
    
    if not collection_urls:
        print("Không có URL nào được nhập. Thoát.")
        exit()
    
    print(f"\nSẽ crawl {len(collection_urls)} collection(s)")
    
    crawler = TheNewOriginalsCrawler(collection_urls)
    crawler.run()
