import os
import re
import time
from playwright.sync_api import sync_playwright
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from dotenv import load_dotenv

load_dotenv()

cloudinary_url = os.getenv('CLOUDINARY_URL')
if cloudinary_url:
    match = re.match(r'cloudinary://([^:]+):([^@]+)@(.+)', cloudinary_url)
    if match:
        api_key, api_secret, cloud_name = match.groups()
        cloudinary.config(
            cloud_name=cloud_name,
            api_key=api_key,
            api_secret=api_secret
        )

class ProductNameFormatter:
    """Format product name theo PRODUCT_NAMING_GUIDE"""
    
    FIT_STYLES = [
        'Relaxed Fit', 'Slim Fit', 'Regular Fit', 'Oversized', 
        'Crop Top', 'Dài Tay', 'Ngắn Tay', 'Ba Lỗ',
        'Bomber', 'Denim', 'Cardigan', 'Zip-up'
    ]
    
    PRODUCT_TYPES = [
        'Áo Thun', 'Áo Sơ Mi', 'Áo Khoác', 'Áo Hoodie', 'Áo Polo',
        'Áo Len', 'Quần Jean', 'Quần Kaki', 'Quần Short'
    ]
    
    @staticmethod
    def extract_product_type(name):
        """Extract loại sản phẩm từ tên"""
        name_upper = name.upper()
        for ptype in ProductNameFormatter.PRODUCT_TYPES:
            if ptype.upper() in name_upper:
                return ptype
        return 'Áo Thun'
    
    @staticmethod
    def extract_fit_style(name):
        """Extract fit/style từ tên"""
        for fit in ProductNameFormatter.FIT_STYLES:
            if fit.lower() in name.lower():
                return fit
        return None
    
    @staticmethod
    def extract_design_name(name, product_type, fit_style):
        """Extract tên design (phần còn lại sau khi bỏ type và fit)"""
        clean_name = name
        
        clean_name = re.sub(r'(?i)cotton\s+cao\s+cấp', '', clean_name)
        clean_name = re.sub(r'(?i)cotton\s+100%?', '', clean_name)
        clean_name = re.sub(r'(?i)chất\s+liệu.*?(?=\s|$)', '', clean_name)
        
        if product_type:
            clean_name = clean_name.replace(product_type, '')
        if fit_style:
            clean_name = clean_name.replace(fit_style, '')
        
        clean_name = re.sub(r'\s+', ' ', clean_name).strip()
        
        return clean_name if clean_name else 'Classic'
    
    @staticmethod
    def format_name(original_name):
        """
        Format: [Loại sản phẩm] [Fit/Style] - [Tên Design]
        VD: "Áo Thun Relaxed Fit - Summer Vibes Sea Life"
        """
        product_type = ProductNameFormatter.extract_product_type(original_name)
        fit_style = ProductNameFormatter.extract_fit_style(original_name)
        design_name = ProductNameFormatter.extract_design_name(original_name, product_type, fit_style)
        
        if fit_style:
            return f"{product_type} {fit_style} - {design_name}"
        else:
            return f"{product_type} - {design_name}"

class DescriptionGenerator:
    """Generate description theo PRODUCT_NAMING_GUIDE template"""
    
    MATERIALS = [
        "cotton 100% mềm mại, thoáng mát, thấm hút mồ hôi tốt",
        "cotton cao cấp mềm mịn, co giãn nhẹ thoải mái",
        "cotton pha spandex thoáng khí, giữ form tốt",
        "vải cotton premium mát mẻ, không xù lông",
    ]
    
    STYLES = [
        "phong cách casual hàng ngày, đi chơi cuối tuần, đi cafe với bạn bè",
        "phong cách streetwear năng động, đi học, đi làm",
        "phong cách minimalist hiện đại, dễ phối đồ",
        "phong cách trẻ trung, năng động, phù hợp mọi dịp",
    ]
    
    @staticmethod
    def generate_intro(product_name, original_desc):
        """Generate phần giới thiệu dựa vào tên và description gốc"""
        design_match = re.search(r'-\s*(.+)$', product_name)
        design_name = design_match.group(1).strip() if design_match else ""
        
        product_type = product_name.split('-')[0].strip().lower()
        
        if 'relaxed fit' in product_name.lower():
            fit_desc = "với form rộng thoải mái"
        elif 'slim fit' in product_name.lower():
            fit_desc = "với form ôm vừa vặn"
        elif 'oversized' in product_name.lower():
            fit_desc = "với form rộng oversized cá tính"
        else:
            fit_desc = "với thiết kế hiện đại"
        
        if design_name and len(design_name) > 3:
            return f"{product_type.capitalize()} {fit_desc}, họa tiết {design_name} độc đáo và bắt mắt"
        else:
            return f"{product_type.capitalize()} {fit_desc}, thiết kế tối giản sang trọng"
    
    @staticmethod
    def generate_keywords(product_name, color_name):
        """Generate keywords cho search"""
        keywords = set()
        
        if 'áo thun' in product_name.lower():
            keywords.update(['áo thun', 't-shirt', 'áo phông'])
        elif 'áo sơ mi' in product_name.lower():
            keywords.update(['áo sơ mi', 'shirt', 'sơ mi'])
        elif 'áo khoác' in product_name.lower():
            keywords.update(['áo khoác', 'jacket', 'khoác'])
        elif 'áo hoodie' in product_name.lower():
            keywords.update(['áo hoodie', 'hoodie', 'áo nỉ'])
        
        if 'relaxed fit' in product_name.lower():
            keywords.add('áo rộng')
        elif 'slim fit' in product_name.lower():
            keywords.add('áo ôm')
        elif 'oversized' in product_name.lower():
            keywords.add('áo oversized')
        
        if color_name and color_name.lower() != 'n/a':
            keywords.add(f'áo {color_name.lower()}')
        
        design_match = re.search(r'-\s*(.+)$', product_name)
        if design_match:
            design = design_match.group(1).strip().lower()
            if len(design) > 3:
                keywords.add(f'áo {design}')
        
        return ', '.join(sorted(keywords))
    
    @staticmethod
    def generate(product_name, color_name, original_desc=""):
        """
        Generate full description theo template:
        [Intro] + [Material] + [Style] + [Keywords]
        """
        import random
        
        intro = DescriptionGenerator.generate_intro(product_name, original_desc)
        material = random.choice(DescriptionGenerator.MATERIALS)
        style = random.choice(DescriptionGenerator.STYLES)
        keywords = DescriptionGenerator.generate_keywords(product_name, color_name)
        
        description = f"{intro}. Chất liệu {material}. Phù hợp cho {style}.\nKeywords: {keywords}."
        
        return description

class ColorParser:
    """Parse và xử lý màu sắc"""
    
    @staticmethod
    def extract_first_color(color_name):
        """
        Nếu màu ghép (VD: "Trắng Cổ Đen"), lấy từ đầu tiên
        """
        if not color_name or color_name.lower() == 'n/a':
            return color_name
        
        parts = re.split(r'\s+', color_name.strip())
        return parts[0]
    
    @staticmethod
    def normalize_color_name(color_name):
        """Chuẩn hóa tên màu: capitalize từ đầu"""
        if not color_name:
            return 'N/A'
        first_color = ColorParser.extract_first_color(color_name)
        return first_color.capitalize()

class PriceParser:
    """Parse giá thành số thuần"""
    
    @staticmethod
    def parse(price_text):
        """
        Convert "159.000 đ" -> 159000
        """
        if not price_text:
            return 0
        
        clean = re.sub(r'[^\d]', '', price_text)
        
        try:
            return int(clean) if clean else 0
        except:
            return 0

class CategoryParser:
    """Parse category name từ URL"""
    
    @staticmethod
    def parse(collection_url):
        """
        Convert "/collections/ao-thun-relaxed-fit" -> "Áo Thun Relaxed Fit"
        """
        match = re.search(r'/collections/([^/?]+)', collection_url)
        if not match:
            return 'Unknown'
        
        slug = match.group(1)
        
        words = slug.split('-')
        capitalized = [word.capitalize() for word in words]
        
        return ' '.join(capitalized)

class SeedDataCrawler:
    def __init__(self, collection_urls):
        self.collection_urls = collection_urls if isinstance(collection_urls, list) else [collection_urls]
        self.excel_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'seed_data.xlsx')
        
        self.categories = {}
        self.colors = {}
        self.products = []
        
        self.category_id_counter = 1
        self.color_id_counter = 1
        self.product_id_counter = 1
        
        self.crawled_products = set()
    
    def upload_to_cloudinary(self, image_url, folder_name, timeout=30):
        try:
            if not image_url.startswith('http'):
                image_url = 'https:' + image_url if image_url.startswith('//') else 'https://theneworiginals.co' + image_url
            
            result = cloudinary.uploader.upload(
                image_url,
                folder=f"theneworiginals/{folder_name}",
                use_filename=True,
                unique_filename=True,
                timeout=timeout
            )
            return result['secure_url']
        except Exception as e:
            print(f"⚠️ Upload failed: {str(e)[:100]}")
            return None
    
    def get_or_create_category(self, category_name):
        """Get category ID, tạo mới nếu chưa có"""
        if category_name in self.categories:
            return self.categories[category_name]
        
        cat_id = self.category_id_counter
        self.categories[category_name] = cat_id
        self.category_id_counter += 1
        return cat_id
    
    def get_or_create_color(self, color_name):
        """Get color ID, tạo mới nếu chưa có"""
        normalized = ColorParser.normalize_color_name(color_name)
        
        if normalized in self.colors:
            return self.colors[normalized]
        
        color_id = self.color_id_counter
        self.colors[normalized] = color_id
        self.color_id_counter += 1
        return color_id
    
    def crawl_collection(self, page, collection_url, max_pages=25, max_products=100):
        print(f"\nCrawling collection: {collection_url} (max {max_products} products, max {max_pages} pages)")
        
        all_products = set()
        current_page = 1
        
        while True:
            if current_page > max_pages:
                print(f"  Reached page limit ({max_pages})")
                break
            
            if len(all_products) >= max_products:
                print(f"  Reached product limit ({max_products})")
                break
            
            page_url = f"{collection_url}?page={current_page}" if current_page > 1 else collection_url
            print(f"  Page {current_page}/{max_pages}...", end=' ')
            
            try:
                page.goto(page_url, wait_until='domcontentloaded', timeout=60000)
                time.sleep(2)
                
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(1)
                
                products_on_page = page.evaluate("""
                    () => {
                        const links = Array.from(document.querySelectorAll('a[href*="/products/"]'));
                        const baseUrls = links.map(a => a.href.split('?')[0]);
                        return [...new Set(baseUrls)];
                    }
                """)
                
                if not products_on_page or len(products_on_page) == 0:
                    print("No products found")
                    break
                
                before_count = len(all_products)
                for product in products_on_page:
                    if len(all_products) >= max_products:
                        break
                    all_products.add(product)
                new_products = len(all_products) - before_count
                
                print(f"{new_products} new products (total: {len(all_products)})")
                
                if len(all_products) >= max_products:
                    print(f"  Reached product limit ({max_products})")
                    break
                
                has_next = page.evaluate("""
                    () => {
                        const nextButton = document.querySelector('.pagination__item--next:not(.pagination__item--disable)');
                        return nextButton !== null;
                    }
                """)
                
                if not has_next:
                    break
                
                current_page += 1
                
            except Exception as e:
                print(f"Error on page {current_page}: {str(e)[:50]}")
                break
        
        product_links = list(all_products)[:max_products]
        print(f"\nTotal unique products: {len(product_links)}")
        return product_links
    
    def get_all_colors(self, page):
        try:
            colors = page.evaluate("""
                () => {
                    const colorNames = [];
                    
                    const colorInputs = document.querySelectorAll('input[name="Màu"]');
                    colorInputs.forEach(input => {
                        const colorValue = input.value;
                        if (colorValue) {
                            colorNames.push(colorValue);
                        }
                    });
                    
                    if (colorNames.length === 0) {
                        const currentColor = document.querySelector('.current-option[data-selected-value]');
                        if (currentColor) {
                            colorNames.push(currentColor.textContent.trim());
                        }
                    }
                    
                    return colorNames;
                }
            """)
            return colors if colors and len(colors) > 0 else ['N/A']
        except:
            return ['N/A']
    
    def crawl_product_detail(self, page, product_url, category_name):
        print(f"\nCrawling product: {product_url}")
        try:
            page.goto(product_url, wait_until='domcontentloaded', timeout=45000)
            time.sleep(2)
        except Exception as e:
            print(f"⚠️ Failed to load product page: {str(e)[:50]}")
            return
        
        original_name = page.evaluate("""
            () => {
                const h1 = document.querySelector('h1.product__title, .description-block__heading');
                return h1?.textContent.trim() || 'Unknown Product';
            }
        """)
        
        if original_name in self.crawled_products:
            print(f"  ⏭️  Skipped (already crawled)")
            return
        
        price_text = page.evaluate("""
            () => {
                const priceEl = document.querySelector('.price, [class*="price"], .product-price');
                return priceEl?.textContent.trim() || '0';
            }
        """)
        
        colors_list = self.get_all_colors(page)
        
        original_desc = page.evaluate("""
            () => {
                const descBlock = document.querySelector('.description-block__text .rte');
                if (descBlock) {
                    return descBlock.textContent.trim();
                }
                return '';
            }
        """)
        
        formatted_name = ProductNameFormatter.format_name(original_name)
        price = PriceParser.parse(price_text)
        
        color_ids = []
        for color in colors_list:
            color_id = self.get_or_create_color(color)
            color_ids.append(color_id)
        
        first_color_name = ColorParser.normalize_color_name(colors_list[0]) if colors_list else 'N/A'
        
        description = DescriptionGenerator.generate(formatted_name, first_color_name, original_desc)
        
        print(f"  Original: {original_name}")
        print(f"  Formatted: {formatted_name}")
        print(f"  Price: {price}")
        print(f"  Colors: {', '.join(colors_list)} -> IDs: {color_ids}")
        
        images = page.evaluate("""
            () => {
                const imgs = [];
                
                const productImages = document.querySelectorAll('.product-image img, .product-gallery img, [class*="ProductImage"] img, .product__media img');
                productImages.forEach(img => {
                    const src = img.src || img.getAttribute('data-src') || img.getAttribute('srcset')?.split(' ')[0];
                    if (src && !src.includes('icon') && !src.includes('logo')) {
                        const cleanSrc = src.split('?')[0];
                        imgs.push(cleanSrc);
                    }
                });
                
                if (imgs.length === 0) {
                    const allImgs = document.querySelectorAll('img');
                    allImgs.forEach(img => {
                        const parent = img.closest('.header, .footer, .nav, nav, .menu');
                        if (!parent) {
                            const src = img.src || img.getAttribute('data-src');
                            if (src && src.includes('theneworiginals') && !src.includes('icon') && !src.includes('logo')) {
                                const cleanSrc = src.split('?')[0];
                                imgs.push(cleanSrc);
                            }
                        }
                    });
                }
                
                return [...new Set(imgs)];
            }
        """)
        
        print(f"  Found {len(images)} images")
        
        uploaded_images = []
        max_images = min(len(images), 10)
        for img_idx, img_url in enumerate(images[:max_images]):
            print(f"  [{img_idx+1}/{max_images}] Uploading...", end=' ')
            uploaded_url = self.upload_to_cloudinary(img_url, f"{category_name}/{formatted_name.replace(' ', '_')}", timeout=30)
            if uploaded_url:
                uploaded_images.append(uploaded_url)
                print("✓")
            else:
                print("✗ Skip")
        
        if len(uploaded_images) > 0:
            category_id = self.get_or_create_category(category_name)
            
            product_data = {
                'id': self.product_id_counter,
                'category_id': category_id,
                'name': formatted_name,
                'description': description,
                'selling_price': price,
                'color_ids': ','.join(map(str, color_ids)),
                'images': ', '.join(uploaded_images)
            }
            
            self.products.append(product_data)
            self.crawled_products.add(original_name)
            self.product_id_counter += 1
            
            print(f"  ✓ Saved product ID={product_data['id']} with {len(uploaded_images)} images")
        else:
            print(f"  ⚠️ No images saved")
    
    def save_to_excel(self):
        """Save data vào Excel với 3 sheets"""
        from datetime import datetime
        
        print(f"\n{'='*60}")
        print("Saving to Excel...")
        
        try:
            wb = Workbook()
            
            ws_categories = wb.active
            ws_categories.title = "Categories"
            ws_categories.append(['id', 'name'])
            for cat_name, cat_id in sorted(self.categories.items(), key=lambda x: x[1]):
                ws_categories.append([cat_id, cat_name])
            
            ws_categories['A1'].font = Font(bold=True)
            ws_categories['B1'].font = Font(bold=True)
            ws_categories.column_dimensions['A'].width = 10
            ws_categories.column_dimensions['B'].width = 30
            
            ws_colors = wb.create_sheet("Colors")
            ws_colors.append(['id', 'name'])
            for color_name, color_id in sorted(self.colors.items(), key=lambda x: x[1]):
                ws_colors.append([color_id, color_name])
            
            ws_colors['A1'].font = Font(bold=True)
            ws_colors['B1'].font = Font(bold=True)
            ws_colors.column_dimensions['A'].width = 10
            ws_colors.column_dimensions['B'].width = 20
            
            ws_products = wb.create_sheet("Products")
            ws_products.append(['id', 'category_id', 'name', 'description', 'selling_price', 'color_ids', 'images'])
            
            for product in self.products:
                ws_products.append([
                    product['id'],
                    product['category_id'],
                    product['name'],
                    product['description'],
                    product['selling_price'],
                    product['color_ids'],
                    product['images']
                ])
            
            for col in ws_products[1]:
                col.font = Font(bold=True)
                col.alignment = Alignment(horizontal='center', vertical='center')
            
            ws_products.column_dimensions['A'].width = 10
            ws_products.column_dimensions['B'].width = 15
            ws_products.column_dimensions['C'].width = 40
            ws_products.column_dimensions['D'].width = 60
            ws_products.column_dimensions['E'].width = 15
            ws_products.column_dimensions['F'].width = 20
            ws_products.column_dimensions['G'].width = 80
            
            wb.save(self.excel_path)
            
            print(f"✓ Excel saved: {self.excel_path}")
            print(f"  - Categories: {len(self.categories)}")
            print(f"  - Colors: {len(self.colors)}")
            print(f"  - Products: {len(self.products)}")
            print(f"{'='*60}")
            
        except Exception as e:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(os.path.expanduser('~'), 'Downloads', f'seed_data_{timestamp}.xlsx')
            print(f"⚠️ Failed to save to {self.excel_path}")
            print(f"   Trying backup: {backup_path}")
            
            wb.save(backup_path)
            self.excel_path = backup_path
            print(f"✓ Saved to backup location")
    
    def run(self):
        print(f"Total collections to crawl: {len(self.collection_urls)}\n")
        
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=False)
                page = browser.new_page()
                
                all_products = []
                for idx, collection_url in enumerate(self.collection_urls, 1):
                    print(f"\n{'='*60}")
                    print(f"[Collection {idx}/{len(self.collection_urls)}]")
                    category_name = CategoryParser.parse(collection_url)
                    print(f"Category: {category_name}")
                    
                    try:
                        product_links = self.crawl_collection(page, collection_url)
                        for product in product_links:
                            all_products.append((product, category_name))
                    except Exception as e:
                        print(f"Error crawling collection {collection_url}: {e}")
                        continue
                
                print(f"\n{'='*60}")
                print(f"Total products found: {len(all_products)}")
                print(f"{'='*60}\n")
                
                for idx, (product_url, category_name) in enumerate(all_products, 1):
                    print(f"\n{'='*60}")
                    print(f"[Product {idx}/{len(all_products)}]")
                    print(f"Progress: {len(self.products)} products saved so far")
                    print(f"URL: {product_url}")
                    
                    try:
                        self.crawl_product_detail(page, product_url, category_name)
                    except KeyboardInterrupt:
                        raise
                    except Exception as e:
                        print(f"⚠️ Error crawling product: {str(e)[:100]}")
                        print("→ Skipping to next product...")
                        continue
                
                browser.close()
        except KeyboardInterrupt:
            print("\n\n" + "="*60)
            print("⚠️ SCRIPT INTERRUPTED BY USER (Ctrl+C)")
            print(f"Data saved: {len(self.products)} products")
            print("="*60)
        except Exception as e:
            print(f"\n\n⚠️ Script error: {e}")
        finally:
            self.save_to_excel()

if __name__ == "__main__":
    print("=== SEED DATA CRAWLER ===\n")
    print("Nhập danh sách collection URLs (mỗi URL 1 dòng, nhấn Enter 2 lần để kết thúc):")
    print("Ví dụ: https://theneworiginals.co/collections/ao-thun-relaxed-fit\n")
    
    collection_urls = []
    while True:
        url = input().strip()
        if not url:
            break
        if '/collections/' in url:
            collection_urls.append(url)
        else:
            print("⚠️  URL không hợp lệ (phải chứa '/collections/')")
    
    if not collection_urls:
        print("Không có URL nào được nhập. Thoát.")
        exit()
    
    print(f"\nSẽ crawl {len(collection_urls)} collection(s)")
    
    crawler = SeedDataCrawler(collection_urls)
    crawler.run()
